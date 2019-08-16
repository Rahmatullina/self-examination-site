from django.http import HttpResponseRedirect, HttpResponseNotFound, HttpResponse,Http404
from django.urls import reverse
from django.shortcuts import render
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from .forms import SE_Form, LoginForm
from .models import RegionModel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

regions_names = ['Абзелиловский район РБ', 'Агидель РБ', 'Альшеевский район РБ', 'Архангельский район РБ',
                 'Аскинский район РБ', 'Аургазинский район РБ', 'Баймакский район РБ', 'Бакалинский район РБ',
                 'Балтачевский район РБ', 'Белебеевский район РБ', 'Белокатайский район РБ', 'Белорецкий район РБ',
                 'Бижбулякский район РБ', 'Бирский район РБ', 'Благоварский район РБ', 'Благовещенский район РБ',
                 'Буздякский район РБ', 'Бураевский район РБ', 'Бурзянский район РБ', 'Гафурийский район РБ',
                 'Давлекановский район РБ', 'Дуванский район РБ', 'Дюртюлинский район РБ', 'Ермекеевский район',
                 'Зианчуринский район РБ', 'Зилаирский район РБ', 'Иглинский район РБ', 'Илишевский район РБ',
                 'Ишимбайский район РБ', 'Калтасинский район РБ', 'Караидельский район РБ', 'Кармаскалинский район РБ',
                 'Кигинский район РБ', 'Краснокамский район РБ', 'Кугарчинский район РБ', 'Кумертау',
                 'Кушнаренковский район РБ', 'Куюргазинский район РБ', 'Межгорье', 'Мелеузовский район РБ',
                 'Мечетлинский район РБ', 'Мишкинский район РБ', 'Миякинский район РБ', 'Нефтекамск',
                 'Нуримановский район РБ', 'Октябрьский', 'Салават', 'Салаватский район РБ',
                 'Сибай', 'Стерлибашевский район РБ', 'Стерлитамак', 'Стерлитамакский район РБ',
                 'Татышлинский район РБ', 'Туймазинский район РБ', 'Уфа', 'Уфимский район РБ', 'Учалинский район РБ',
                 'Федоровский район РБ', 'Хайбуллинский район РБ', 'Чекмагушевский район РБ', 'Чишминский район РБ',
                 'Шаранский район РБ', 'Янаульский район РБ']

short_regions_names = ['abzelil', 'agidel', 'alsheev', 'archang', 'askinsk', 'aurgazin', 'baymak', 'bakalin', 'baltach',
                       'belebeev', 'belokatay', 'belorezk', 'bizhbul', 'birsk', 'blagovar', 'blagovesch', 'buzdyak',
                       'buraev', 'burzyan', 'gafur', 'davlekan', 'duvansk', 'dyurtyulin', 'ermekeev', 'zianchurin',
                       'zilairsk', 'iglinsk', 'ilishevsk', 'ishimbaysk', 'kaltasinsk', 'karaidelsks', 'karmaskalin',
                       'kiginsk', 'krasnokamsk', 'kugarchinsk', 'kumertau', 'kushnarenk', 'kuyurgazinsk', 'mezhgorie',
                       'meleuz', 'mechetlin', 'mishkin', 'miyakin', 'neftekamsk', 'nurimanovsk', 'oktabrsks',
                       'slavat', 'salavatskiy', 'sibi', 'sterlibash', 'sterlitamak', 'sterlitamakskiy',
                       'tatyshlin', 'tuymazin', 'ufa', 'ufimsk', 'uchalinsk', 'fedorovsk', 'haybullinsk',
                       'chekmagush', 'chishminsk', 'sharanck', 'yanaulsk']

short_service_names = ['residential_premises','housing_transfer','advertising_structures','capital_construction',
                       'preschool_education','school_education','needing_premises','town_planning','archive_reference','land_schemes']

MONTHS = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

MONTH_NUMBERS = ['01','02','03','04','05','06','07','08','09','10','11','12']

class dotDict(dict):
    """dot.notation access to dictionary attributes"""
    __getattr__ = dict.get
    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__

def get_not_sent(month,year):
    objects = []
    for name in regions_names:
        try:
            obj = RegionModel.objects.raw('''
                    SELECT id
                    FROM app_regionModel 
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' +
                                          str(month) + '''\' AND year=\'''' + str(year) + '''\'''')[0]
        except IndexError:
            objects.append(name)

    return objects


def get_with_troubles(month, year):
    objects = dict()
    objects.update({'residential_premises': []})
    objects.update({'housing_transfer': []})
    objects.update({'advertising_structures': []})
    objects.update({'capital_construction': []})
    objects.update({'preschool_education': []})
    objects.update({'school_education': []})
    objects.update({'needing_premises': []})
    objects.update({'town_planning': []})
    objects.update({'archive_reference': []})
    objects.update({'land_schemes': []})
    for name in regions_names:
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    residential_premises_id_RGMU,
                    residential_premises_statement_amount,
                    residential_premises_link,
                    residential_premises_has_advanced_appointment_comment,
                    residential_premises_has_btn_get_service_comment,
                    residential_premises_has_reglament_comment,
                    residential_premises_has_estimation_quality_comment,
                    residential_premises_connected_to_FGIS_DO_comment,
                    residential_premises_has_electronic_form_printing_comment,
                    residential_premises_has_edition_draft_comment,
                    residential_premises_has_term_of_consideration_comment,
                    residential_premises_has_notif_consider_result_comment,
                    residential_premises_has_causes_of_failure_comment,
                    residential_premises_has_sample_document_comment,
                    residential_premises_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.residential_premises_has_advanced_appointment_comment != 'Да' and
                obj.residential_premises_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_btn_get_service_comment != 'Да' and
                     obj.residential_premises_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_reglament_comment != 'Да' and
                     obj.residential_premises_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_estimation_quality_comment != 'Да' and
                     obj.residential_premises_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_connected_to_FGIS_DO_comment != 'Да' and
                     obj.residential_premises_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_electronic_form_printing_comment != 'Да' and
                     obj.residential_premises_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_edition_draft_comment != 'Да' and
                     obj.residential_premises_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_term_of_consideration_comment != 'Да' and
                     obj.residential_premises_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_notif_consider_result_comment != 'Да' and
                     obj.residential_premises_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_causes_of_failure_comment != 'Да' and
                     obj.residential_premises_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_sample_document_comment != 'Да' and
                     obj.residential_premises_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.residential_premises_has_document_template_comment != 'Да' and
                     obj.residential_premises_has_document_template_comment != 'Не предусмотрено'):
                objects['residential_premises'].append(obj)

        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    housing_transfer_id_RGMU,
                    housing_transfer_statement_amount,
                    housing_transfer_link,
                    housing_transfer_has_advanced_appointment_comment,
                    housing_transfer_has_btn_get_service_comment,
                    housing_transfer_has_reglament_comment,
                    housing_transfer_has_estimation_quality_comment,
                    housing_transfer_connected_to_FGIS_DO_comment,
                    housing_transfer_has_electronic_form_printing_comment,
                    housing_transfer_has_edition_draft_comment,
                    housing_transfer_has_term_of_consideration_comment,
                    housing_transfer_has_notif_consider_result_comment,
                    housing_transfer_has_causes_of_failure_comment,
                    housing_transfer_has_sample_document_comment,
                    housing_transfer_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.housing_transfer_has_advanced_appointment_comment != 'Да' and
                obj.housing_transfer_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_btn_get_service_comment != 'Да' and
                     obj.housing_transfer_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_reglament_comment != 'Да' and
                     obj.housing_transfer_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_estimation_quality_comment != 'Да' and
                     obj.housing_transfer_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_connected_to_FGIS_DO_comment != 'Да' and
                     obj.housing_transfer_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_electronic_form_printing_comment != 'Да' and
                     obj.housing_transfer_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_edition_draft_comment != 'Да' and
                     obj.housing_transfer_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_term_of_consideration_comment != 'Да' and
                     obj.housing_transfer_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_notif_consider_result_comment != 'Да' and
                     obj.housing_transfer_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_causes_of_failure_comment != 'Да' and
                     obj.housing_transfer_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_sample_document_comment != 'Да' and
                     obj.housing_transfer_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.housing_transfer_has_document_template_comment != 'Да' and
                     obj.housing_transfer_has_document_template_comment != 'Не предусмотрено'):
                objects['housing_transfer'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                   advertising_structures_id_RGMU,
                   advertising_structures_statement_amount,
                   advertising_structures_link,
                   advertising_structures_has_advanced_appointment_comment,
                   advertising_structures_has_btn_get_service_comment,
                   advertising_structures_has_reglament_comment,
                   advertising_structures_has_estimation_quality_comment,
                   advertising_structures_connected_to_FGIS_DO_comment,
                   advertising_structures_has_electronic_form_printing_comment,
                   advertising_structures_has_edition_draft_comment,
                   advertising_structures_has_term_of_consideration_comment,
                   advertising_structures_has_notif_consider_result_comment,
                   advertising_structures_has_causes_of_failure_comment,
                   advertising_structures_has_sample_document_comment,
                   advertising_structures_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.advertising_structures_has_advanced_appointment_comment != 'Да' and
                obj.advertising_structures_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_btn_get_service_comment != 'Да' and
                     obj.advertising_structures_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_reglament_comment != 'Да' and
                     obj.advertising_structures_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_estimation_quality_comment != 'Да' and
                     obj.advertising_structures_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_connected_to_FGIS_DO_comment != 'Да' and
                     obj.advertising_structures_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_electronic_form_printing_comment != 'Да' and
                     obj.advertising_structures_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_edition_draft_comment != 'Да' and
                     obj.advertising_structures_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_term_of_consideration_comment != 'Да' and
                     obj.advertising_structures_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_notif_consider_result_comment != 'Да' and
                     obj.advertising_structures_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_causes_of_failure_comment != 'Да' and
                     obj.advertising_structures_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_sample_document_comment != 'Да' and
                     obj.advertising_structures_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.advertising_structures_has_document_template_comment != 'Да' and
                     obj.advertising_structures_has_document_template_comment != 'Не предусмотрено'):
                objects['advertising_structures'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    capital_construction_id_RGMU,
                    capital_construction_statement_amount,
                    capital_construction_link,
                    capital_construction_has_advanced_appointment_comment,
                    capital_construction_has_btn_get_service_comment,
                    capital_construction_has_reglament_comment,
                    capital_construction_has_estimation_quality_comment,
                    capital_construction_connected_to_FGIS_DO_comment,
                    capital_construction_has_electronic_form_printing_comment,
                    capital_construction_has_edition_draft_comment,
                    capital_construction_has_term_of_consideration_comment,
                    capital_construction_has_notif_consider_result_comment,
                    capital_construction_has_causes_of_failure_comment,
                    capital_construction_has_sample_document_comment,
                    capital_construction_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.capital_construction_has_advanced_appointment_comment != 'Да' and
                obj.capital_construction_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_btn_get_service_comment != 'Да' and
                     obj.capital_construction_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_reglament_comment != 'Да' and
                     obj.capital_construction_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_estimation_quality_comment != 'Да' and
                     obj.capital_construction_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_connected_to_FGIS_DO_comment != 'Да' and
                     obj.capital_construction_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_electronic_form_printing_comment != 'Да' and
                     obj.capital_construction_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_edition_draft_comment != 'Да' and
                     obj.capital_construction_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_term_of_consideration_comment != 'Да' and
                     obj.capital_construction_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_notif_consider_result_comment != 'Да' and
                     obj.capital_construction_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_causes_of_failure_comment != 'Да' and
                     obj.capital_construction_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_sample_document_comment != 'Да' and
                     obj.capital_construction_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.capital_construction_has_document_template_comment != 'Да' and
                     obj.capital_construction_has_document_template_comment != 'Не предусмотрено'):
                objects['capital_construction'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    preschool_education_id_RGMU,
                    preschool_education_statement_amount,
                    preschool_education_link,
                    preschool_education_has_advanced_appointment_comment,
                    preschool_education_has_btn_get_service_comment,
                    preschool_education_has_reglament_comment,
                    preschool_education_has_estimation_quality_comment,
                    preschool_education_connected_to_FGIS_DO_comment,
                    preschool_education_has_electronic_form_printing_comment,
                    preschool_education_has_edition_draft_comment,
                    preschool_education_has_term_of_consideration_comment,
                    preschool_education_has_notif_consider_result_comment,
                    preschool_education_has_causes_of_failure_comment,
                    preschool_education_has_sample_document_comment,
                    preschool_education_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.preschool_education_has_advanced_appointment_comment != 'Да' and
                obj.preschool_education_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_btn_get_service_comment != 'Да' and
                     obj.preschool_education_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_reglament_comment != 'Да' and
                     obj.preschool_education_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_estimation_quality_comment != 'Да' and
                     obj.preschool_education_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_connected_to_FGIS_DO_comment != 'Да' and
                     obj.preschool_education_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_electronic_form_printing_comment != 'Да' and
                     obj.preschool_education_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_edition_draft_comment != 'Да' and
                     obj.preschool_education_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_term_of_consideration_comment != 'Да' and
                     obj.preschool_education_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_notif_consider_result_comment != 'Да' and
                     obj.preschool_education_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_causes_of_failure_comment != 'Да' and
                     obj.preschool_education_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_sample_document_comment != 'Да' and
                     obj.preschool_education_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.preschool_education_has_document_template_comment != 'Да' and
                     obj.preschool_education_has_document_template_comment != 'Не предусмотрено'):
                objects['preschool_education'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    residential_premises_id_RGMU,
                    school_education_statement_amount,
                    school_education_link,
                    school_education_has_advanced_appointment_comment,
                    school_education_has_btn_get_service_comment,
                    school_education_has_reglament_comment,
                    school_education_has_estimation_quality_comment,
                    school_education_connected_to_FGIS_DO_comment,
                    school_education_has_electronic_form_printing_comment,
                    school_education_has_edition_draft_comment,
                    school_education_has_term_of_consideration_comment,
                    school_education_has_notif_consider_result_comment,
                    school_education_has_causes_of_failure_comment,
                    school_education_has_sample_document_comment,
                    school_education_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.school_education_has_advanced_appointment_comment != 'Да' and
                obj.school_education_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_btn_get_service_comment != 'Да' and
                     obj.school_education_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_reglament_comment != 'Да' and
                     obj.school_education_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_estimation_quality_comment != 'Да' and
                     obj.school_education_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.school_education_connected_to_FGIS_DO_comment != 'Да' and
                     obj.school_education_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_electronic_form_printing_comment != 'Да' and
                     obj.school_education_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_edition_draft_comment != 'Да' and
                     obj.school_education_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_term_of_consideration_comment != 'Да' and
                     obj.school_education_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_notif_consider_result_comment != 'Да' and
                     obj.school_education_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_causes_of_failure_comment != 'Да' and
                     obj.school_education_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_sample_document_comment != 'Да' and
                     obj.school_education_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.school_education_has_document_template_comment != 'Да' and
                     obj.school_education_has_document_template_comment != 'Не предусмотрено'):
                objects['school_education'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    needing_premises_id_RGMU,
                    needing_premises_statement_amount,
                    needing_premises_link,
                    needing_premises_has_advanced_appointment_comment,
                    needing_premises_has_btn_get_service_comment,
                    needing_premises_has_reglament_comment,
                    needing_premises_has_estimation_quality_comment,
                    needing_premises_connected_to_FGIS_DO_comment,
                    needing_premises_has_electronic_form_printing_comment,
                    needing_premises_has_edition_draft_comment,
                    needing_premises_has_term_of_consideration_comment,
                    needing_premises_has_notif_consider_result_comment,
                    needing_premises_has_causes_of_failure_comment,
                    needing_premises_has_sample_document_comment,
                    needing_premises_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.needing_premises_has_advanced_appointment_comment != 'Да' and
                obj.needing_premises_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_btn_get_service_comment != 'Да' and
                     obj.needing_premises_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_reglament_comment != 'Да' and
                     obj.needing_premises_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_estimation_quality_comment != 'Да' and
                     obj.needing_premises_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_connected_to_FGIS_DO_comment != 'Да' and
                     obj.needing_premises_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_electronic_form_printing_comment != 'Да' and
                     obj.needing_premises_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_edition_draft_comment != 'Да' and
                     obj.needing_premises_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_term_of_consideration_comment != 'Да' and
                     obj.needing_premises_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_notif_consider_result_comment != 'Да' and
                     obj.needing_premises_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_causes_of_failure_comment != 'Да' and
                     obj.needing_premises_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_sample_document_comment != 'Да' and
                     obj.needing_premises_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.needing_premises_has_document_template_comment != 'Да' and
                     obj.needing_premises_has_document_template_comment != 'Не предусмотрено'):
                objects['needing_premises'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    town_planning_id_RGMU,
                    town_planning_statement_amount,
                    town_planning_link,
                    town_planning_has_advanced_appointment_comment,
                    town_planning_has_btn_get_service_comment,
                    town_planning_has_reglament_comment,
                    town_planning_has_estimation_quality_comment,
                    town_planning_connected_to_FGIS_DO_comment,
                    town_planning_has_electronic_form_printing_comment,
                    town_planning_has_edition_draft_comment,
                    town_planning_has_term_of_consideration_comment,
                    town_planning_has_notif_consider_result_comment,
                    town_planning_has_causes_of_failure_comment,
                    town_planning_has_sample_document_comment,
                    town_planning_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.town_planning_has_advanced_appointment_comment != 'Да' and
                obj.town_planning_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_btn_get_service_comment != 'Да' and
                     obj.town_planning_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_reglament_comment != 'Да' and
                     obj.town_planning_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_estimation_quality_comment != 'Да' and
                     obj.town_planning_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.town_planning_connected_to_FGIS_DO_comment != 'Да' and
                     obj.town_planning_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_electronic_form_printing_comment != 'Да' and
                     obj.town_planning_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_edition_draft_comment != 'Да' and
                     obj.town_planning_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_term_of_consideration_comment != 'Да' and
                     obj.town_planning_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_notif_consider_result_comment != 'Да' and
                     obj.town_planning_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_causes_of_failure_comment != 'Да' and
                     obj.town_planning_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_sample_document_comment != 'Да' and
                     obj.town_planning_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.town_planning_has_document_template_comment != 'Да' and
                     obj.town_planning_has_document_template_comment != 'Не предусмотрено'):
                objects['town_planning'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    archive_reference_id_RGMU,
                    archive_reference_statement_amount,
                    archive_reference_link,
                    archive_reference_has_advanced_appointment_comment,
                    archive_reference_has_btn_get_service_comment,
                    archive_reference_has_reglament_comment,
                    archive_reference_has_estimation_quality_comment,
                    archive_reference_connected_to_FGIS_DO_comment,
                    archive_reference_has_electronic_form_printing_comment,
                    archive_reference_has_edition_draft_comment,
                    archive_reference_has_term_of_consideration_comment,
                    archive_reference_has_notif_consider_result_comment,
                    archive_reference_has_causes_of_failure_comment,
                    archive_reference_has_sample_document_comment,
                    archive_reference_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.archive_reference_has_advanced_appointment_comment != 'Да' and
                obj.archive_reference_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_btn_get_service_comment != 'Да' and
                     obj.archive_reference_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_reglament_comment != 'Да' and
                     obj.archive_reference_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_estimation_quality_comment != 'Да' and
                     obj.archive_reference_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_connected_to_FGIS_DO_comment != 'Да' and
                     obj.archive_reference_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_electronic_form_printing_comment != 'Да' and
                     obj.archive_reference_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_edition_draft_comment != 'Да' and
                     obj.archive_reference_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_term_of_consideration_comment != 'Да' and
                     obj.archive_reference_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_notif_consider_result_comment != 'Да' and
                     obj.archive_reference_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_causes_of_failure_comment != 'Да' and
                     obj.archive_reference_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_sample_document_comment != 'Да' and
                     obj.archive_reference_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.archive_reference_has_document_template_comment != 'Да' and
                     obj.archive_reference_has_document_template_comment != 'Не предусмотрено'):
                objects['archive_reference'].append(obj)
        for obj in RegionModel.objects.raw(
                '''SELECT id, region_name,
                    land_schemes_id_RGMU,
                    land_schemes_statement_amount,
                    land_schemes_link,
                    land_schemes_has_advanced_appointment_comment,
                    land_schemes_has_btn_get_service_comment,
                    land_schemes_has_reglament_comment,
                    land_schemes_has_estimation_quality_comment,
                    land_schemes_connected_to_FGIS_DO_comment,
                    land_schemes_has_electronic_form_printing_comment,
                    land_schemes_has_edition_draft_comment,
                    land_schemes_has_term_of_consideration_comment,
                    land_schemes_has_notif_consider_result_comment,
                    land_schemes_has_causes_of_failure_comment,
                    land_schemes_has_sample_document_comment,
                    land_schemes_has_document_template_comment\
                    FROM app_regionModel WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(month) +
                '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;'''):
            if (obj.land_schemes_has_advanced_appointment_comment != 'Да' and
                obj.land_schemes_has_advanced_appointment_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_btn_get_service_comment != 'Да' and
                     obj.land_schemes_has_btn_get_service_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_reglament_comment != 'Да' and
                     obj.land_schemes_has_reglament_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_estimation_quality_comment != 'Да' and
                     obj.land_schemes_has_estimation_quality_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_connected_to_FGIS_DO_comment != 'Да' and
                     obj.land_schemes_connected_to_FGIS_DO_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_electronic_form_printing_comment != 'Да' and
                     obj.land_schemes_has_electronic_form_printing_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_edition_draft_comment != 'Да' and
                     obj.land_schemes_has_edition_draft_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_term_of_consideration_comment != 'Да' and
                     obj.land_schemes_has_term_of_consideration_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_notif_consider_result_comment != 'Да' and
                     obj.land_schemes_has_notif_consider_result_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_causes_of_failure_comment != 'Да' and
                     obj.land_schemes_has_causes_of_failure_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_sample_document_comment != 'Да' and
                     obj.land_schemes_has_sample_document_comment != 'Не предусмотрено') or \
                    (obj.land_schemes_has_document_template_comment != 'Да' and
                     obj.land_schemes_has_document_template_comment != 'Не предусмотрено'):
                objects['land_schemes'].append(obj)
    return objects

@login_required
def get_self_examination_form(request):
    """
    View function for renewing a specific regionForm by users
    """
    if request.user.is_authenticated:
        print(request.user.username)
        if request.method == 'POST':

            # Create a form instance and populate it with data from the request (binding):
            region_form = SE_Form(request.POST)

            # Check if the form is valid:
            if region_form.is_valid():

                form_to_save = region_form.save(commit=False)
                form_to_save.time = datetime.today().time()
                form_to_save.day = datetime.today().strftime('%d')
                form_to_save.year = datetime.today().strftime('%Y')
                form_to_save.month = datetime.today().strftime('%m')
                form_to_save.region_name = request.user.region_name
                form_to_save.save()

                # redirect to a new URL:
                return HttpResponseRedirect(reverse('result_form', kwargs={
                    'service_name': 'residential_premises',
                    'year': datetime.today().strftime('%Y'),
                    'month': datetime.today().strftime('%m'),
                }))

        # If this is a GET (or any other method) create the default form.
        else:
            region_form = SE_Form(initial={
                                       'residential_premises_id_RGMU': 'id_GRMU_redevelop1',
                                       'residential_premises_statement_amount': '555',
                                       'residential_premises_link': 'http:\\my_link_1.com',
                                       'housing_transfer_id_RGMU': 'id_GRMU_transfer1',
                                       'housing_transfer_statement_amount': '55566788',
                                       'housing_transfer_link': 'http:\\my_link_1234.com',
                                       'advertising_structures_id_RGMU': 'id_GRMU_advert1',
                                       'advertising_structures_statement_amount': '55566788',
                                       'advertising_structures_link': 'http:\\my_link_1234.com',
                                       'capital_construction_id_RGMU': 'id_GRMU_capital1',
                                       'capital_construction_statement_amount': '55566788',
                                       'capital_construction_link': 'http:\\my_link_1234.com',
                                       'preschool_education_id_RGMU': 'id_GRMU_preschool1',
                                       'preschool_education_statement_amount': '55566788',
                                       'preschool_education_link': 'http:\\my_link_1234.com',
                                       'school_education_id_RGMU': 'id_GRMU_school1',
                                       'school_education_statement_amount': '55566788',
                                       'school_education_link': 'http:\\my_link_1234.com',
                                       'needing_premises_id_RGMU': 'id_GRMU_residental1',
                                       'needing_premises_statement_amount': '55566788',
                                       'needing_premises_link': 'http:\\my_link_1234.com',
                                       'town_planning_id_RGMU': 'id_GRMU_town1',
                                       'town_planning_statement_amount': '55566788',
                                       'town_planning_link': 'http:\\my_link_1234.com',
                                       'archive_reference_id_RGMU': 'id_GRMU_archive1',
                                       'archive_reference_statement_amount': '55566788',
                                       'archive_reference_link': 'http:\\my_link_1234.com',
                                       'land_schemes_id_RGMU': 'id_GRMU_land1',
                                       'land_schemes_statement_amount': '55566788',
                                       'land_schemes_link': 'http:\\my_link_1234.com',
                                       'month': str(datetime.today().strftime('%m')),
                                       'year': str(datetime.today().strftime('%Y')),
                                       })

        return render(request, 'app/self_examination_form.html', {
            'region_form': region_form,
            'username': request.user.username,
            'region_name': request.user.region_name,
            'year':datetime.today().strftime('%Y'),
            'num_month':datetime.today().strftime('%m'),
            'zipped': zip(regions_names, short_regions_names),
        })
    else:
        return HttpResponseRedirect(reverse('login'))


@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def get_result_form(request, service_name, year, month):
    objects = list()

    if service_name == 'residential_premises':
        full_service_name = 'Прием заявлений и выдача документов о согласовании проведения переустройства и (или) ' \
                            'перепланировки жилого помещения '
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw('''SELECT id, region_name,
                    residential_premises_id_RGMU AS id_RGMU,
                    residential_premises_statement_amount AS statement_amount,
                    residential_premises_link AS link,
                    residential_premises_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    residential_premises_has_btn_get_service_comment AS has_btn_get_service_comment,
                    residential_premises_has_reglament_comment AS has_reglament_comment,
                    residential_premises_has_estimation_quality_comment AS has_estimation_quality_comment,
                    residential_premises_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    residential_premises_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    residential_premises_has_edition_draft_comment AS has_edition_draft_comment,
                    residential_premises_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    residential_premises_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    residential_premises_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    residential_premises_has_sample_document_comment AS has_sample_document_comment,
                    residential_premises_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                    month) + '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)

            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'housing_transfer':
        full_service_name = 'Принятие решений о переводе жилых помещений в нежилые помещения и нежилых помещений в ' \
                            'жилые помещения '
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    housing_transfer_id_RGMU AS id_RGMU,
                    housing_transfer_statement_amount AS statement_amount,
                    housing_transfer_link AS link,
                    housing_transfer_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    housing_transfer_has_btn_get_service_comment AS has_btn_get_service_comment,
                    housing_transfer_has_reglament_comment AS has_reglament_comment,
                    housing_transfer_has_estimation_quality_comment AS has_estimation_quality_comment,
                    housing_transfer_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    housing_transfer_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    housing_transfer_has_edition_draft_comment AS has_edition_draft_comment,
                    housing_transfer_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    housing_transfer_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    housing_transfer_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    housing_transfer_has_sample_document_comment AS has_sample_document_comment,
                    housing_transfer_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'advertising_structures':
        full_service_name = 'Выдача разрешения на установку и эксплуатацию рекламной конструкции'
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    advertising_structures_id_RGMU AS id_RGMU,
                    advertising_structures_statement_amount AS statement_amount,
                    advertising_structures_link AS link,
                    advertising_structures_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    advertising_structures_has_btn_get_service_comment AS has_btn_get_service_comment,
                    advertising_structures_has_reglament_comment AS has_reglament_comment,
                    advertising_structures_has_estimation_quality_comment AS has_estimation_quality_comment,
                    advertising_structures_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    advertising_structures_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    advertising_structures_has_edition_draft_comment AS has_edition_draft_comment,
                    advertising_structures_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    advertising_structures_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    advertising_structures_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    advertising_structures_has_sample_document_comment AS has_sample_document_comment,
                    advertising_structures_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'capital_construction':
        full_service_name = 'Выдача разрешения на строительство, реконструкцию объектов капитального строительства'
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    capital_construction_id_RGMU AS id_RGMU,
                    capital_construction_statement_amount AS statement_amount,
                    capital_construction_link AS link,
                    capital_construction_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    capital_construction_has_btn_get_service_comment AS has_btn_get_service_comment,
                    capital_construction_has_reglament_comment AS has_reglament_comment,
                    capital_construction_has_estimation_quality_comment AS has_estimation_quality_comment,
                    capital_construction_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    capital_construction_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    capital_construction_has_edition_draft_comment AS has_edition_draft_comment,
                    capital_construction_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    capital_construction_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    capital_construction_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    capital_construction_has_sample_document_comment AS has_sample_document_comment,
                    capital_construction_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'preschool_education':
        full_service_name = 'Прием заявлений, постановка на учет и зачисление детей в образовательные учреждения, ' \
                            'реализующие основную образовательную программу дошкольного образования (детские сады) '
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    preschool_education_id_RGMU AS id_RGMU,
                    preschool_education_statement_amount AS statement_amount,
                    preschool_education_link AS link,
                    preschool_education_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    preschool_education_has_btn_get_service_comment AS has_btn_get_service_comment,
                    preschool_education_has_reglament_comment AS has_reglament_comment,
                    preschool_education_has_estimation_quality_comment AS has_estimation_quality_comment,
                    preschool_education_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    preschool_education_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    preschool_education_has_edition_draft_comment AS has_edition_draft_comment,
                    preschool_education_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    preschool_education_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    preschool_education_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    preschool_education_has_sample_document_comment AS has_sample_document_comment,
                    preschool_education_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'school_education':
        full_service_name = 'Зачисление детей в общеобразовательные учреждения субъектов Российской Федерации или ' \
                            'муниципальные общеобразовательные учреждения '
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    school_education_id_RGMU AS id_RGMU,
                    school_education_statement_amount AS statement_amount,
                    school_education_link AS link,
                    school_education_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    school_education_has_btn_get_service_comment AS has_btn_get_service_comment,
                    school_education_has_reglament_comment AS has_reglament_comment,
                    school_education_has_estimation_quality_comment AS has_estimation_quality_comment,
                    school_education_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    school_education_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    school_education_has_edition_draft_comment AS has_edition_draft_comment,
                    school_education_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    school_education_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    school_education_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    school_education_has_sample_document_comment AS has_sample_document_comment,
                    school_education_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'needing_premises':
        full_service_name = 'Принятие на учет граждан в качестве нуждающихся в жилых помещениях'
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    needing_premises_id_RGMU AS id_RGMU,
                    needing_premises_statement_amount AS statement_amount,
                    needing_premises_link AS link,
                    needing_premises_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    needing_premises_has_btn_get_service_comment AS has_btn_get_service_comment,
                    needing_premises_has_reglament_comment AS has_reglament_comment,
                    needing_premises_has_estimation_quality_comment AS has_estimation_quality_comment,
                    needing_premises_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    needing_premises_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    needing_premises_has_edition_draft_comment AS has_edition_draft_comment,
                    needing_premises_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    needing_premises_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    needing_premises_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    needing_premises_has_sample_document_comment AS has_sample_document_comment,
                    needing_premises_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'town_planning':
        full_service_name = 'Выдача градостроительных планов земельных участков'
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    town_planning_id_RGMU AS id_RGMU,
                    town_planning_statement_amount AS statement_amount,
                    town_planning_link AS link,
                    town_planning_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    town_planning_has_btn_get_service_comment AS has_btn_get_service_comment,
                    town_planning_has_reglament_comment AS has_reglament_comment,
                    town_planning_has_estimation_quality_comment AS has_estimation_quality_comment,
                    town_planning_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    town_planning_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    town_planning_has_edition_draft_comment AS has_edition_draft_comment,
                    town_planning_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    town_planning_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    town_planning_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    town_planning_has_sample_document_comment AS has_sample_document_comment,
                    town_planning_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'archive_reference':
        full_service_name = 'Предоставление архивных справок, архивных копий, архивных выписок, информационных писем, ' \
                            'связанных с реализацией законных прав и свобод граждан и исполнением государственными ' \
                            'органами и органами местного самоуправления своих полномочий '
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    archive_reference_id_RGMU AS id_RGMU,
                    archive_reference_statement_amount AS statement_amount,
                    archive_reference_link AS link,
                    archive_reference_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    archive_reference_has_btn_get_service_comment AS has_btn_get_service_comment,
                    archive_reference_has_reglament_comment AS has_reglament_comment,
                    archive_reference_has_estimation_quality_comment AS has_estimation_quality_comment,
                    archive_reference_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    archive_reference_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    archive_reference_has_edition_draft_comment AS has_edition_draft_comment,
                    archive_reference_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    archive_reference_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    archive_reference_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    archive_reference_has_sample_document_comment AS has_sample_document_comment,
                    archive_reference_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'land_schemes':
        full_service_name = 'Утверждение схемы расположения земельного участка или земельных участков на кадастровом ' \
                            'плане территории '
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    land_schemes_id_RGMU AS id_RGMU,
                    land_schemes_statement_amount AS statement_amount,
                    land_schemes_link AS link,
                    land_schemes_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    land_schemes_has_btn_get_service_comment AS has_btn_get_service_comment,
                    land_schemes_has_reglament_comment AS has_reglament_comment,
                    land_schemes_has_estimation_quality_comment AS has_estimation_quality_comment,
                    land_schemes_connected_to_FGIS_DO_comment AS connected_to_FGIS_DO_comment,
                    land_schemes_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    land_schemes_has_edition_draft_comment AS has_edition_draft_comment,
                    land_schemes_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    land_schemes_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    land_schemes_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    land_schemes_has_sample_document_comment AS has_sample_document_comment,
                    land_schemes_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_RGMU': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_FGIS_DO_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    else:
        return HttpResponseNotFound('')

    return render(request, 'app/result_form.html', {'objects': objects,
                                                        'year': str(year),
                                                        'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                        'num_month': month,
                                                        'full_service_name': full_service_name,
                                                        'service_name': service_name,
                                                        'zipped': zip(regions_names, short_regions_names),
                                                        'username': request.user.username,
                                                        'years': [i for i in range(2016, int(datetime.now().year)+1)]})


@login_required(login_url='/login/',
                redirect_field_name='/result_form/' + datetime.today().strftime('%Y/%m/')+'ufa/')
def get_region_form(request, year, month, short_region_name):
    full_region_name = regions_names[short_regions_names.index(short_region_name)]
    try:
        object = RegionModel.objects.raw(
            '''SELECT id,
            residential_premises_id_RGMU,
            residential_premises_statement_amount, 
            residential_premises_link,
            residential_premises_has_advanced_appointment_comment,
            residential_premises_has_btn_get_service_comment,
            residential_premises_has_reglament_comment,
            residential_premises_has_estimation_quality_comment,
            residential_premises_connected_to_FGIS_DO_comment,
            residential_premises_has_electronic_form_printing_comment,
            residential_premises_has_edition_draft_comment,
            residential_premises_has_term_of_consideration_comment,
            residential_premises_has_notif_consider_result_comment,
            residential_premises_has_causes_of_failure_comment,
            residential_premises_has_sample_document_comment,
            residential_premises_has_document_template_comment,
            housing_transfer_id_RGMU,
            housing_transfer_statement_amount, 
            housing_transfer_link,
            housing_transfer_has_advanced_appointment_comment,
            housing_transfer_has_btn_get_service_comment,
            housing_transfer_has_reglament_comment,
            housing_transfer_has_estimation_quality_comment,
            housing_transfer_connected_to_FGIS_DO_comment,
            housing_transfer_has_electronic_form_printing_comment,
            housing_transfer_has_edition_draft_comment,
            housing_transfer_has_term_of_consideration_comment,
            housing_transfer_has_notif_consider_result_comment,
            housing_transfer_has_causes_of_failure_comment,
            housing_transfer_has_sample_document_comment,
            housing_transfer_has_document_template_comment,
            advertising_structures_id_RGMU,
            advertising_structures_statement_amount, 
            advertising_structures_link,
            advertising_structures_has_advanced_appointment_comment,
            advertising_structures_has_btn_get_service_comment,
            advertising_structures_has_reglament_comment,
            advertising_structures_has_estimation_quality_comment,
            advertising_structures_connected_to_FGIS_DO_comment,
            advertising_structures_has_electronic_form_printing_comment,
            advertising_structures_has_edition_draft_comment,
            advertising_structures_has_term_of_consideration_comment,
            advertising_structures_has_notif_consider_result_comment,
            advertising_structures_has_causes_of_failure_comment,
            advertising_structures_has_sample_document_comment,
            advertising_structures_has_document_template_comment,
            capital_construction_id_RGMU,
            capital_construction_statement_amount, 
            capital_construction_link,
            capital_construction_has_advanced_appointment_comment,
            capital_construction_has_btn_get_service_comment,
            capital_construction_has_reglament_comment,
            capital_construction_has_estimation_quality_comment,
            capital_construction_connected_to_FGIS_DO_comment,
            capital_construction_has_electronic_form_printing_comment,
            capital_construction_has_edition_draft_comment,
            capital_construction_has_term_of_consideration_comment,
            capital_construction_has_notif_consider_result_comment,
            capital_construction_has_causes_of_failure_comment,
            capital_construction_has_sample_document_comment,
            capital_construction_has_document_template_comment,
            preschool_education_id_RGMU,
            preschool_education_statement_amount, 
            preschool_education_link,
            preschool_education_has_advanced_appointment_comment,
            preschool_education_has_btn_get_service_comment,
            preschool_education_has_reglament_comment,
            preschool_education_has_estimation_quality_comment,
            preschool_education_connected_to_FGIS_DO_comment,
            preschool_education_has_electronic_form_printing_comment,
            preschool_education_has_edition_draft_comment,
            preschool_education_has_term_of_consideration_comment,
            preschool_education_has_notif_consider_result_comment,
            preschool_education_has_causes_of_failure_comment,
            preschool_education_has_sample_document_comment,
            preschool_education_has_document_template_comment,
            school_education_id_RGMU,
            school_education_statement_amount, 
            school_education_link,
            school_education_has_advanced_appointment_comment,
            school_education_has_btn_get_service_comment,
            school_education_has_reglament_comment,
            school_education_has_estimation_quality_comment,
            school_education_connected_to_FGIS_DO_comment,
            school_education_has_electronic_form_printing_comment,
            school_education_has_edition_draft_comment,
            school_education_has_term_of_consideration_comment,
            school_education_has_notif_consider_result_comment,
            school_education_has_causes_of_failure_comment,
            school_education_has_sample_document_comment,
            school_education_has_document_template_comment,
            needing_premises_id_RGMU,
            needing_premises_statement_amount, 
            needing_premises_link,
            needing_premises_has_advanced_appointment_comment,
            needing_premises_has_btn_get_service_comment,
            needing_premises_has_reglament_comment,
            needing_premises_has_estimation_quality_comment,
            needing_premises_connected_to_FGIS_DO_comment,
            needing_premises_has_electronic_form_printing_comment,
            needing_premises_has_edition_draft_comment,
            needing_premises_has_term_of_consideration_comment,
            needing_premises_has_notif_consider_result_comment,
            needing_premises_has_causes_of_failure_comment,
            needing_premises_has_sample_document_comment,
            needing_premises_has_document_template_comment,
            town_planning_id_RGMU,
            town_planning_statement_amount, 
            town_planning_link,
            town_planning_has_advanced_appointment_comment,
            town_planning_has_btn_get_service_comment,
            town_planning_has_reglament_comment,
            town_planning_has_estimation_quality_comment,
            town_planning_connected_to_FGIS_DO_comment,
            town_planning_has_electronic_form_printing_comment,
            town_planning_has_edition_draft_comment,
            town_planning_has_term_of_consideration_comment,
            town_planning_has_notif_consider_result_comment,
            town_planning_has_causes_of_failure_comment,
            town_planning_has_sample_document_comment,
            town_planning_has_document_template_comment,
            archive_reference_id_RGMU,
            archive_reference_statement_amount, 
            archive_reference_link,
            archive_reference_has_advanced_appointment_comment,
            archive_reference_has_btn_get_service_comment,
            archive_reference_has_reglament_comment,
            archive_reference_has_estimation_quality_comment,
            archive_reference_connected_to_FGIS_DO_comment,
            archive_reference_has_electronic_form_printing_comment,
            archive_reference_has_edition_draft_comment,
            archive_reference_has_term_of_consideration_comment,
            archive_reference_has_notif_consider_result_comment,
            archive_reference_has_causes_of_failure_comment,
            archive_reference_has_sample_document_comment,
            archive_reference_has_document_template_comment,
            land_schemes_id_RGMU,
            land_schemes_statement_amount, 
            land_schemes_link,
            land_schemes_has_advanced_appointment_comment,
            land_schemes_has_btn_get_service_comment,
            land_schemes_has_reglament_comment,
            land_schemes_has_estimation_quality_comment,
            land_schemes_connected_to_FGIS_DO_comment,
            land_schemes_has_electronic_form_printing_comment,
            land_schemes_has_edition_draft_comment,
            land_schemes_has_term_of_consideration_comment,
            land_schemes_has_notif_consider_result_comment,
            land_schemes_has_causes_of_failure_comment,
            land_schemes_has_sample_document_comment,
            land_schemes_has_document_template_comment
            FROM app_regionModel
            WHERE region_name =\'''' + str(full_region_name) + '''\' AND month=\'''' + str(
                month) + '''\' AND year=\'''' + str(
                year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]

    except IndexError:
        object = {'short_region_name': short_region_name,
                  'residential_premises_id_RGMU': "",
                  'residential_premises_statement_amount': "",
                  'residential_premises_link': "",
                  'residential_premises_has_advanced_appointment_comment': "",
                  'residential_premises_has_btn_get_service_comment': "",
                  'residential_premises_has_reglament_comment': "",
                  'residential_premises_has_estimation_quality_comment': "",
                  'residential_premises_connected_to_FGIS_DO_comment': "",
                  'residential_premises_has_electronic_form_printing_comment': "",
                  'residential_premises_has_edition_draft_comment': "",
                  'residential_premises_has_term_of_consideration_comment': "",
                  'residential_premises_has_notif_consider_result_comment': "",
                  'residential_premises_has_causes_of_failure_comment': "",
                  'residential_premises_has_sample_document_comment': "",
                  'residential_premises_has_document_template_comment': "",
                  'housing_transfer_id_RGMU': "",
                  'housing_transfer_statement_amount': "",
                  'housing_transfer_link': "",
                  'housing_transfer_has_advanced_appointment_comment': "",
                  'housing_transfer_has_btn_get_service_comment': "",
                  'housing_transfer_has_reglament_comment': "",
                  'housing_transfer_has_estimation_quality_comment': "",
                  'housing_transfer_connected_to_FGIS_DO_comment': "",
                  'housing_transfer_has_electronic_form_printing_comment': "",
                  'housing_transfer_has_edition_draft_comment': "",
                  'housing_transfer_has_term_of_consideration_comment': "",
                  'housing_transfer_has_notif_consider_result_comment': "",
                  'housing_transfer_has_causes_of_failure_comment': "",
                  'housing_transfer_has_sample_document_comment': "",
                  'housing_transfer_has_document_template_comment': "",
                  'advertising_structures_id_RGMU': "",
                  'advertising_structures_statement_amount': "",
                  'advertising_structures_link': "",
                  'advertising_structures_has_advanced_appointment_comment': "",
                  'advertising_structures_has_btn_get_service_comment': "",
                  'advertising_structures_has_reglament_comment': "",
                  'advertising_structures_has_estimation_quality_comment': "",
                  'advertising_structures_connected_to_FGIS_DO_comment': "",
                  'advertising_structures_has_electronic_form_printing_comment': "",
                  'advertising_structures_has_edition_draft_comment': "",
                  'advertising_structures_has_term_of_consideration_comment': "",
                  'advertising_structures_has_notif_consider_result_comment': "",
                  'advertising_structures_has_causes_of_failure_comment': "",
                  'advertising_structures_has_sample_document_comment': "",
                  'advertising_structures_has_document_template_comment': "",
                  'capital_construction_id_RGMU': "",
                  'capital_construction_statement_amount': "",
                  'capital_construction_link': "",
                  'capital_construction_has_advanced_appointment_comment': "",
                  'capital_construction_has_btn_get_service_comment': "",
                  'capital_construction_has_reglament_comment': "",
                  'capital_construction_has_estimation_quality_comment': "",
                  'capital_construction_connected_to_FGIS_DO_comment': "",
                  'capital_construction_has_electronic_form_printing_comment': "",
                  'capital_construction_has_edition_draft_comment': "",
                  'capital_construction_has_term_of_consideration_comment': "",
                  'capital_construction_has_notif_consider_result_comment': "",
                  'capital_construction_has_causes_of_failure_comment': "",
                  'capital_construction_has_sample_document_comment': "",
                  'capital_construction_has_document_template_comment': "",
                  'preschool_education_id_RGMU': "",
                  'preschool_education_statement_amount': "",
                  'preschool_education_link': "",
                  'preschool_education_has_advanced_appointment_comment': "",
                  'preschool_education_has_btn_get_service_comment': "",
                  'preschool_education_has_reglament_comment': "",
                  'preschool_education_has_estimation_quality_comment': "",
                  'preschool_education_connected_to_FGIS_DO_comment': "",
                  'preschool_education_has_electronic_form_printing_comment': "",
                  'preschool_education_has_edition_draft_comment': "",
                  'preschool_education_has_term_of_consideration_comment': "",
                  'preschool_education_has_notif_consider_result_comment': "",
                  'preschool_education_has_causes_of_failure_comment': "",
                  'preschool_education_has_sample_document_comment': "",
                  'preschool_education_has_document_template_comment': "",
                  'school_education_id_RGMU': "",
                  'school_education_statement_amount': "",
                  'school_education_link': "",
                  'school_education_has_advanced_appointment_comment': "",
                  'school_education_has_btn_get_service_comment': "",
                  'school_education_has_reglament_comment': "",
                  'school_education_has_estimation_quality_comment': "",
                  'school_education_connected_to_FGIS_DO_comment': "",
                  'school_education_has_electronic_form_printing_comment': "",
                  'school_education_has_edition_draft_comment': "",
                  'school_education_has_term_of_consideration_comment': "",
                  'school_education_has_notif_consider_result_comment': "",
                  'school_education_has_causes_of_failure_comment': "",
                  'school_education_has_sample_document_comment': "",
                  'school_education_has_document_template_comment': "",
                  'needing_premises_id_RGMU': "",
                  'needing_premises_statement_amount': "",
                  'needing_premises_link': "",
                  'needing_premises_has_advanced_appointment_comment': "",
                  'needing_premises_has_btn_get_service_comment': "",
                  'needing_premises_has_reglament_comment': "",
                  'needing_premises_has_estimation_quality_comment': "",
                  'needing_premises_connected_to_FGIS_DO_comment': "",
                  'needing_premises_has_electronic_form_printing_comment': "",
                  'needing_premises_has_edition_draft_comment': "",
                  'needing_premises_has_term_of_consideration_comment': "",
                  'needing_premises_has_notif_consider_result_comment': "",
                  'needing_premises_has_causes_of_failure_comment': "",
                  'needing_premises_has_sample_document_comment': "",
                  'needing_premises_has_document_template_comment': "",
                  'town_planning_id_RGMU': "",
                  'town_planning_statement_amount': "",
                  'town_planning_link': "",
                  'town_planning_has_advanced_appointment_comment': "",
                  'town_planning_has_btn_get_service_comment': "",
                  'town_planning_has_reglament_comment': "",
                  'town_planning_has_estimation_quality_comment': "",
                  'town_planning_connected_to_FGIS_DO_comment': "",
                  'town_planning_has_electronic_form_printing_comment': "",
                  'town_planning_has_edition_draft_comment': "",
                  'town_planning_has_term_of_consideration_comment': "",
                  'town_planning_has_notif_consider_result_comment': "",
                  'town_planning_has_causes_of_failure_comment': "",
                  'town_planning_has_sample_document_comment': "",
                  'town_planning_has_document_template_comment': "",
                  'archive_reference_id_RGMU': "",
                  'archive_reference_statement_amount': "",
                  'archive_reference_link': "",
                  'archive_reference_has_advanced_appointment_comment': "",
                  'archive_reference_has_btn_get_service_comment': "",
                  'archive_reference_has_reglament_comment': "",
                  'archive_reference_has_estimation_quality_comment': "",
                  'archive_reference_connected_to_FGIS_DO_comment': "",
                  'archive_reference_has_electronic_form_printing_comment': "",
                  'archive_reference_has_edition_draft_comment': "",
                  'archive_reference_has_term_of_consideration_comment': "",
                  'archive_reference_has_notif_consider_result_comment': "",
                  'archive_reference_has_causes_of_failure_comment': "",
                  'archive_reference_has_sample_document_comment': "",
                  'archive_reference_has_document_template_comment': "",
                  'land_schemes_id_RGMU': "",
                  'land_schemes_statement_amount': "",
                  'land_schemes_link': "",
                  'land_schemes_has_advanced_appointment_comment': "",
                  'land_schemes_has_btn_get_service_comment': "",
                  'land_schemes_has_reglament_comment': "",
                  'land_schemes_has_estimation_quality_comment': "",
                  'land_schemes_connected_to_FGIS_DO_comment': "",
                  'land_schemes_has_electronic_form_printing_comment': "",
                  'land_schemes_has_edition_draft_comment': "",
                  'land_schemes_has_term_of_consideration_comment': "",
                  'land_schemes_has_notif_consider_result_comment': "",
                  'land_schemes_has_causes_of_failure_comment': "",
                  'land_schemes_has_sample_document_comment': "",
                  'land_schemes_has_document_template_comment': ""
                  }
    return render(request, 'app/region_form.html', {'object': object,
                                                        'year': str(year),
                                                        'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                        'num_month': month,
                                                        'zipped': zip(regions_names, short_regions_names),
                                                        'full_region_name': full_region_name,
                                                        'short_region_name': short_region_name,
                                                        'username': request.user.username,
                                                        'years': [i for i in range(2016, int(datetime.now().year) + 1)]
                                                        })


@login_required(login_url='/login/',
                redirect_field_name='/result_form/with_troubles/' + datetime.today().strftime('%Y/%m/'))
def get_result_form_with_troubles(request, year, month):
    objects = get_with_troubles(month,year)
    return render(request, 'app/with_troubles.html', {'objects': dotDict(objects),
                                                        'year': str(year),
                                                        'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                        'num_month': month,
                                                        'zipped': zip(regions_names, short_regions_names),
                                                        'username': request.user.username,
                                                        'years': [i for i in range(2016, int(datetime.now().year) + 1)]})


@login_required(login_url='/login/',
                redirect_field_name='/result_form/not_sent/' + datetime.today().strftime('%Y/%m/'))
def get_result_form_not_sent(request, year, month):
    objects = get_not_sent(month,year)
    return render(request, 'app/not_sent.html', {'objects': objects,
                                                     'year': str(year),
                                                     'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                     'num_month': month,
                                                     'zipped': zip(regions_names, short_regions_names),
                                                     'username': request.user.username,
                                                     'years': [i for i in range(2016, int(datetime.now().year) + 1)]
                                                     })


def login_view(request):
    form = LoginForm(request.POST or None)
    print("form created", request.POST,form.is_valid())
    if request.POST and form.is_valid():
        user = form.login(request)
        if user:
            login(request, user)
            return HttpResponseRedirect(reverse('result_form', kwargs={
                'service_name': 'residential_premises',
                'year': datetime.today().strftime('%Y'),
                'month': datetime.today().strftime('%m')
            }))
    return render(request, 'app/registration/login.html', {'form': form})


def logout_view(request):
   logout(request)
   return HttpResponseRedirect('/login/')

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def export_all(request,year,month):
    if request.method == 'GET':
        wb = load_workbook('./app/static/app/all_data.xlsx')

        ws1 = wb["Лист 1"]
        ws2 = wb["Лист 2"]
        ws3 = wb["Лист 3"]
        ws4 = wb["Лист 4"]
        ws5 = wb["Лист 5"]
        ws6 = wb["Лист 6"]
        ws7 = wb["Лист 7"]
        ws8 = wb["Лист 8"]
        ws9 = wb["Лист 9"]
        ws10 = wb["Лист 10"]
        objects = []
        for name in regions_names:
            try:
                object = RegionModel.objects.raw('''SELECT * FROM app_regionModel WHERE region_name =\'''' +
                                                 str(name) + '''\' AND month=\'''' + str(month) + '''\' AND year=\'''' +
                                                 str(year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1''')[0]
                objects.append(object)
            except IndexError:
                object = {'region_name': name,
                          'residential_premises_id_RGMU': "",
                          'residential_premises_statement_amount': "",
                          'residential_premises_link': "",
                          'residential_premises_has_advanced_appointment_comment': "",
                          'residential_premises_has_btn_get_service_comment': "",
                          'residential_premises_has_reglament_comment': "",
                          'residential_premises_has_estimation_quality_comment': "",
                          'residential_premises_connected_to_FGIS_DO_comment': "",
                          'residential_premises_has_electronic_form_printing_comment': "",
                          'residential_premises_has_edition_draft_comment': "",
                          'residential_premises_has_term_of_consideration_comment': "",
                          'residential_premises_has_notif_consider_result_comment': "",
                          'residential_premises_has_causes_of_failure_comment': "",
                          'residential_premises_has_sample_document_comment': "",
                          'residential_premises_has_document_template_comment': "",
                          'housing_transfer_id_RGMU': "",
                          'housing_transfer_statement_amount': "",
                          'housing_transfer_link': "",
                          'housing_transfer_has_advanced_appointment_comment': "",
                          'housing_transfer_has_btn_get_service_comment': "",
                          'housing_transfer_has_reglament_comment': "",
                          'housing_transfer_has_estimation_quality_comment': "",
                          'housing_transfer_connected_to_FGIS_DO_comment': "",
                          'housing_transfer_has_electronic_form_printing_comment': "",
                          'housing_transfer_has_edition_draft_comment': "",
                          'housing_transfer_has_term_of_consideration_comment': "",
                          'housing_transfer_has_notif_consider_result_comment': "",
                          'housing_transfer_has_causes_of_failure_comment': "",
                          'housing_transfer_has_sample_document_comment': "",
                          'housing_transfer_has_document_template_comment': "",
                          'advertising_structures_id_RGMU': "",
                          'advertising_structures_statement_amount': "",
                          'advertising_structures_link': "",
                          'advertising_structures_has_advanced_appointment_comment': "",
                          'advertising_structures_has_btn_get_service_comment': "",
                          'advertising_structures_has_reglament_comment': "",
                          'advertising_structures_has_estimation_quality_comment': "",
                          'advertising_structures_connected_to_FGIS_DO_comment': "",
                          'advertising_structures_has_electronic_form_printing_comment': "",
                          'advertising_structures_has_edition_draft_comment': "",
                          'advertising_structures_has_term_of_consideration_comment': "",
                          'advertising_structures_has_notif_consider_result_comment': "",
                          'advertising_structures_has_causes_of_failure_comment': "",
                          'advertising_structures_has_sample_document_comment': "",
                          'advertising_structures_has_document_template_comment': "",
                          'capital_construction_id_RGMU': "",
                          'capital_construction_statement_amount': "",
                          'capital_construction_link': "",
                          'capital_construction_has_advanced_appointment_comment': "",
                          'capital_construction_has_btn_get_service_comment': "",
                          'capital_construction_has_reglament_comment': "",
                          'capital_construction_has_estimation_quality_comment': "",
                          'capital_construction_connected_to_FGIS_DO_comment': "",
                          'capital_construction_has_electronic_form_printing_comment': "",
                          'capital_construction_has_edition_draft_comment': "",
                          'capital_construction_has_term_of_consideration_comment': "",
                          'capital_construction_has_notif_consider_result_comment': "",
                          'capital_construction_has_causes_of_failure_comment': "",
                          'capital_construction_has_sample_document_comment': "",
                          'capital_construction_has_document_template_comment': "",
                          'preschool_education_id_RGMU': "",
                          'preschool_education_statement_amount': "",
                          'preschool_education_link': "",
                          'preschool_education_has_advanced_appointment_comment': "",
                          'preschool_education_has_btn_get_service_comment': "",
                          'preschool_education_has_reglament_comment': "",
                          'preschool_education_has_estimation_quality_comment': "",
                          'preschool_education_connected_to_FGIS_DO_comment': "",
                          'preschool_education_has_electronic_form_printing_comment': "",
                          'preschool_education_has_edition_draft_comment': "",
                          'preschool_education_has_term_of_consideration_comment': "",
                          'preschool_education_has_notif_consider_result_comment': "",
                          'preschool_education_has_causes_of_failure_comment': "",
                          'preschool_education_has_sample_document_comment': "",
                          'preschool_education_has_document_template_comment': "",
                          'school_education_id_RGMU': "",
                          'school_education_statement_amount': "",
                          'school_education_link': "",
                          'school_education_has_advanced_appointment_comment': "",
                          'school_education_has_btn_get_service_comment': "",
                          'school_education_has_reglament_comment': "",
                          'school_education_has_estimation_quality_comment': "",
                          'school_education_connected_to_FGIS_DO_comment': "",
                          'school_education_has_electronic_form_printing_comment': "",
                          'school_education_has_edition_draft_comment': "",
                          'school_education_has_term_of_consideration_comment': "",
                          'school_education_has_notif_consider_result_comment': "",
                          'school_education_has_causes_of_failure_comment': "",
                          'school_education_has_sample_document_comment': "",
                          'school_education_has_document_template_comment': "",
                          'needing_premises_id_RGMU': "",
                          'needing_premises_statement_amount': "",
                          'needing_premises_link': "",
                          'needing_premises_has_advanced_appointment_comment': "",
                          'needing_premises_has_btn_get_service_comment': "",
                          'needing_premises_has_reglament_comment': "",
                          'needing_premises_has_estimation_quality_comment': "",
                          'needing_premises_connected_to_FGIS_DO_comment': "",
                          'needing_premises_has_electronic_form_printing_comment': "",
                          'needing_premises_has_edition_draft_comment': "",
                          'needing_premises_has_term_of_consideration_comment': "",
                          'needing_premises_has_notif_consider_result_comment': "",
                          'needing_premises_has_causes_of_failure_comment': "",
                          'needing_premises_has_sample_document_comment': "",
                          'needing_premises_has_document_template_comment': "",
                          'town_planning_id_RGMU': "",
                          'town_planning_statement_amount': "",
                          'town_planning_link': "",
                          'town_planning_has_advanced_appointment_comment': "",
                          'town_planning_has_btn_get_service_comment': "",
                          'town_planning_has_reglament_comment': "",
                          'town_planning_has_estimation_quality_comment': "",
                          'town_planning_connected_to_FGIS_DO_comment': "",
                          'town_planning_has_electronic_form_printing_comment': "",
                          'town_planning_has_edition_draft_comment': "",
                          'town_planning_has_term_of_consideration_comment': "",
                          'town_planning_has_notif_consider_result_comment': "",
                          'town_planning_has_causes_of_failure_comment': "",
                          'town_planning_has_sample_document_comment': "",
                          'town_planning_has_document_template_comment': "",
                          'archive_reference_id_RGMU': "",
                          'archive_reference_statement_amount': "",
                          'archive_reference_link': "",
                          'archive_reference_has_advanced_appointment_comment': "",
                          'archive_reference_has_btn_get_service_comment': "",
                          'archive_reference_has_reglament_comment': "",
                          'archive_reference_has_estimation_quality_comment': "",
                          'archive_reference_connected_to_FGIS_DO_comment': "",
                          'archive_reference_has_electronic_form_printing_comment': "",
                          'archive_reference_has_edition_draft_comment': "",
                          'archive_reference_has_term_of_consideration_comment': "",
                          'archive_reference_has_notif_consider_result_comment': "",
                          'archive_reference_has_causes_of_failure_comment': "",
                          'archive_reference_has_sample_document_comment': "",
                          'archive_reference_has_document_template_comment': "",
                          'land_schemes_id_RGMU': "",
                          'land_schemes_statement_amount': "",
                          'land_schemes_link': "",
                          'land_schemes_has_advanced_appointment_comment': "",
                          'land_schemes_has_btn_get_service_comment': "",
                          'land_schemes_has_reglament_comment': "",
                          'land_schemes_has_estimation_quality_comment': "",
                          'land_schemes_connected_to_FGIS_DO_comment': "",
                          'land_schemes_has_electronic_form_printing_comment': "",
                          'land_schemes_has_edition_draft_comment': "",
                          'land_schemes_has_term_of_consideration_comment': "",
                          'land_schemes_has_notif_consider_result_comment': "",
                          'land_schemes_has_causes_of_failure_comment': "",
                          'land_schemes_has_sample_document_comment': "",
                          'land_schemes_has_document_template_comment': ""
                          }
                objects.append(dotDict(object))
        for i in range(len(objects)):
                object = objects[i]
                ws1['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(month, year)
                ws1['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Прием заявлений и выдача документов о согласовании проведения переустройства и (или) перепланировки жилого помещения " в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month,year)
                ws1['C' + str(7+i)] = object.residential_premises_id_RGMU
                ws1['D' + str(7+i)] = object.residential_premises_statement_amount
                ws1['E' + str(7+i)] = object.residential_premises_link
                ws1['F' + str(7+i)] = object.residential_premises_has_advanced_appointment_comment
                ws1['G' + str(7+i)] = object.residential_premises_has_btn_get_service_comment
                ws1['H' + str(7+i)] = object.residential_premises_has_reglament_comment
                ws1['I' + str(7+i)] = object.residential_premises_has_estimation_quality_comment
                ws1['J' + str(7+i)] = object.residential_premises_connected_to_FGIS_DO_comment
                ws1['K' + str(7+i)] = object.residential_premises_has_electronic_form_printing_comment
                ws1['L' + str(7+i)] = object.residential_premises_has_edition_draft_comment
                ws1['M' + str(7+i)] = object.residential_premises_has_term_of_consideration_comment
                ws1['N' + str(7+i)] = object.residential_premises_has_notif_consider_result_comment
                ws1['O' + str(7+i)] = object.residential_premises_has_causes_of_failure_comment
                ws1['P' + str(7+i)] = object.residential_premises_has_sample_document_comment
                ws1['Q' + str(7+i)] = object.residential_premises_has_document_template_comment
                ws2['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(month, year)
                ws2['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Принятие решений о переводе жилых помещений в нежилые помещения и нежилых помещений в жилые помещения" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws2['C' + str(7 + i)] = object.housing_transfer_id_RGMU
                ws2['D' + str(7 + i)] = object.housing_transfer_statement_amount
                ws2['E' + str(7 + i)] = object.housing_transfer_link
                ws2['F' + str(7 + i)] = object.housing_transfer_has_advanced_appointment_comment
                ws2['G' + str(7 + i)] = object.housing_transfer_has_btn_get_service_comment
                ws2['H' + str(7 + i)] = object.housing_transfer_has_reglament_comment
                ws2['I' + str(7 + i)] = object.housing_transfer_has_estimation_quality_comment
                ws2['J' + str(7 + i)] = object.housing_transfer_connected_to_FGIS_DO_comment
                ws2['K' + str(7 + i)] = object.housing_transfer_has_electronic_form_printing_comment
                ws2['L' + str(7 + i)] = object.housing_transfer_has_edition_draft_comment
                ws2['M' + str(7 + i)] = object.housing_transfer_has_term_of_consideration_comment
                ws2['N' + str(7 + i)] = object.housing_transfer_has_notif_consider_result_comment
                ws2['O' + str(7 + i)] = object.housing_transfer_has_causes_of_failure_comment
                ws2['P' + str(7 + i)] = object.housing_transfer_has_sample_document_comment
                ws2['Q' + str(7 + i)] = object.housing_transfer_has_document_template_comment
                ws3['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(month, year)
                ws3['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Выдача разрешения на установку и эксплуатацию рекламной конструкции" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws3['C' + str(7 + i)] = object.advertising_structures_id_RGMU
                ws3['D' + str(7 + i)] = object.advertising_structures_statement_amount
                ws3['E' + str(7 + i)] = object.advertising_structures_link
                ws3['F' + str(7 + i)] = object.advertising_structures_has_advanced_appointment_comment
                ws3['G' + str(7 + i)] = object.advertising_structures_has_btn_get_service_comment
                ws3['H' + str(7 + i)] = object.advertising_structures_has_reglament_comment
                ws3['I' + str(7 + i)] = object.advertising_structures_has_estimation_quality_comment
                ws3['J' + str(7 + i)] = object.advertising_structures_connected_to_FGIS_DO_comment
                ws3['K' + str(7 + i)] = object.advertising_structures_has_electronic_form_printing_comment
                ws3['L' + str(7 + i)] = object.advertising_structures_has_edition_draft_comment
                ws3['M' + str(7 + i)] = object.advertising_structures_has_term_of_consideration_comment
                ws3['N' + str(7 + i)] = object.advertising_structures_has_notif_consider_result_comment
                ws3['O' + str(7 + i)] = object.advertising_structures_has_causes_of_failure_comment
                ws3['P' + str(7 + i)] = object.advertising_structures_has_sample_document_comment
                ws3['Q' + str(7 + i)] = object.advertising_structures_has_document_template_comment
                ws4['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(month, year)
                ws4['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Выдача разрешения на строительство, реконструкцию объектов капитального строительства" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws4['C' + str(7 + i)] = object.capital_construction_id_RGMU
                ws4['D' + str(7 + i)] = object.capital_construction_statement_amount
                ws4['E' + str(7 + i)] = object.capital_construction_link
                ws4['F' + str(7 + i)] = object.capital_construction_has_advanced_appointment_comment
                ws4['G' + str(7 + i)] = object.capital_construction_has_btn_get_service_comment
                ws4['H' + str(7 + i)] = object.capital_construction_has_reglament_comment
                ws4['I' + str(7 + i)] = object.capital_construction_has_estimation_quality_comment
                ws4['J' + str(7 + i)] = object.capital_construction_connected_to_FGIS_DO_comment
                ws4['K' + str(7 + i)] = object.capital_construction_has_electronic_form_printing_comment
                ws4['L' + str(7 + i)] = object.capital_construction_has_edition_draft_comment
                ws4['M' + str(7 + i)] = object.capital_construction_has_term_of_consideration_comment
                ws4['N' + str(7 + i)] = object.capital_construction_has_notif_consider_result_comment
                ws4['O' + str(7 + i)] = object.capital_construction_has_causes_of_failure_comment
                ws4['P' + str(7 + i)] = object.capital_construction_has_sample_document_comment
                ws4['Q' + str(7 + i)] = object.capital_construction_has_document_template_comment
                ws5['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(month, year)
                ws5['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Прием заявлений, постановка на учет и зачисление детей в образовательные учреждения, реализующие основную образовательную программу дошкольного образования (детские сады)" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws5['C' + str(7 + i)] = object.preschool_education_id_RGMU
                ws5['D' + str(7 + i)] = object.preschool_education_statement_amount
                ws5['E' + str(7 + i)] = object.preschool_education_link
                ws5['F' + str(7 + i)] = object.preschool_education_has_advanced_appointment_comment
                ws5['G' + str(7 + i)] = object.preschool_education_has_btn_get_service_comment
                ws5['H' + str(7 + i)] = object.preschool_education_has_reglament_comment
                ws5['I' + str(7 + i)] = object.preschool_education_has_estimation_quality_comment
                ws5['J' + str(7 + i)] = object.preschool_education_connected_to_FGIS_DO_comment
                ws5['K' + str(7 + i)] = object.preschool_education_has_electronic_form_printing_comment
                ws5['L' + str(7 + i)] = object.preschool_education_has_edition_draft_comment
                ws5['M' + str(7 + i)] = object.preschool_education_has_term_of_consideration_comment
                ws5['N' + str(7 + i)] = object.preschool_education_has_notif_consider_result_comment
                ws5['O' + str(7 + i)] = object.preschool_education_has_causes_of_failure_comment
                ws5['P' + str(7 + i)] = object.preschool_education_has_sample_document_comment
                ws5['Q' + str(7 + i)] = object.preschool_education_has_document_template_comment
                ws6['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(month, year)
                ws6['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Зачисление детей в общеобразовательные учреждения субъектов Российской Федерации или муниципальные общеобразовательные учреждения" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws6['C' + str(7 + i)] = object.school_education_id_RGMU
                ws6['D' + str(7 + i)] = object.school_education_statement_amount
                ws6['E' + str(7 + i)] = object.school_education_link
                ws6['F' + str(7 + i)] = object.school_education_has_advanced_appointment_comment
                ws6['G' + str(7 + i)] = object.school_education_has_btn_get_service_comment
                ws6['H' + str(7 + i)] = object.school_education_has_reglament_comment
                ws6['I' + str(7 + i)] = object.school_education_has_estimation_quality_comment
                ws6['J' + str(7 + i)] = object.school_education_connected_to_FGIS_DO_comment
                ws6['K' + str(7 + i)] = object.school_education_has_electronic_form_printing_comment
                ws6['L' + str(7 + i)] = object.school_education_has_edition_draft_comment
                ws6['M' + str(7 + i)] = object.school_education_has_term_of_consideration_comment
                ws6['N' + str(7 + i)] = object.school_education_has_notif_consider_result_comment
                ws6['O' + str(7 + i)] = object.school_education_has_causes_of_failure_comment
                ws6['P' + str(7 + i)] = object.school_education_has_sample_document_comment
                ws6['Q' + str(7 + i)] = object.school_education_has_document_template_comment
                ws7['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(month, year)
                ws7['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Принятие на учет граждан в качестве нуждающихся в жилых помещениях" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws7['C' + str(7 + i)] = object.needing_premises_id_RGMU
                ws7['D' + str(7 + i)] = object.needing_premises_statement_amount
                ws7['E' + str(7 + i)] = object.needing_premises_link
                ws7['F' + str(7 + i)] = object.needing_premises_has_advanced_appointment_comment
                ws7['G' + str(7 + i)] = object.needing_premises_has_btn_get_service_comment
                ws7['H' + str(7 + i)] = object.needing_premises_has_reglament_comment
                ws7['I' + str(7 + i)] = object.needing_premises_has_estimation_quality_comment
                ws7['J' + str(7 + i)] = object.needing_premises_connected_to_FGIS_DO_comment
                ws7['K' + str(7 + i)] = object.needing_premises_has_electronic_form_printing_comment
                ws7['L' + str(7 + i)] = object.needing_premises_has_edition_draft_comment
                ws7['M' + str(7 + i)] = object.needing_premises_has_term_of_consideration_comment
                ws7['N' + str(7 + i)] = object.needing_premises_has_notif_consider_result_comment
                ws7['O' + str(7 + i)] = object.needing_premises_has_causes_of_failure_comment
                ws7['P' + str(7 + i)] = object.needing_premises_has_sample_document_comment
                ws7['Q' + str(7 + i)] = object.needing_premises_has_document_template_comment
                ws8['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(month, year)
                ws8['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Выдача градостроительных планов земельных участков" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws8['C' + str(7 + i)] = object.town_planning_id_RGMU
                ws8['D' + str(7 + i)] = object.town_planning_statement_amount
                ws8['E' + str(7 + i)] = object.town_planning_link
                ws8['F' + str(7 + i)] = object.town_planning_has_advanced_appointment_comment
                ws8['G' + str(7 + i)] = object.town_planning_has_btn_get_service_comment
                ws8['H' + str(7 + i)] = object.town_planning_has_reglament_comment
                ws8['I' + str(7 + i)] = object.town_planning_has_estimation_quality_comment
                ws8['J' + str(7 + i)] = object.town_planning_connected_to_FGIS_DO_comment
                ws8['K' + str(7 + i)] = object.town_planning_has_electronic_form_printing_comment
                ws8['L' + str(7 + i)] = object.town_planning_has_edition_draft_comment
                ws8['M' + str(7 + i)] = object.town_planning_has_term_of_consideration_comment
                ws8['N' + str(7 + i)] = object.town_planning_has_notif_consider_result_comment
                ws8['O' + str(7 + i)] = object.town_planning_has_causes_of_failure_comment
                ws8['P' + str(7 + i)] = object.town_planning_has_sample_document_comment
                ws8['Q' + str(7 + i)] = object.town_planning_has_document_template_comment
                ws9['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в (количество указывается нарастающим итогом)'.format(month, year)
                ws9['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Предоставление архивных справок, архивных копий, архивных выписок, информационных писем, связанных с реализацией законных прав и свобод граждан и исполнением государственными органами и органами местного самоуправления своих полномочий" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws9['C' + str(7 + i)] = object.archive_reference_id_RGMU
                ws9['D' + str(7 + i)] = object.archive_reference_statement_amount
                ws9['E' + str(7 + i)] = object.archive_reference_link
                ws9['F' + str(7 + i)] = object.archive_reference_has_advanced_appointment_comment
                ws9['G' + str(7 + i)] = object.archive_reference_has_btn_get_service_comment
                ws9['H' + str(7 + i)] = object.archive_reference_has_reglament_comment
                ws9['I' + str(7 + i)] = object.archive_reference_has_estimation_quality_comment
                ws9['J' + str(7 + i)] = object.archive_reference_connected_to_FGIS_DO_comment
                ws9['K' + str(7 + i)] = object.archive_reference_has_electronic_form_printing_comment
                ws9['L' + str(7 + i)] = object.archive_reference_has_edition_draft_comment
                ws9['M' + str(7 + i)] = object.archive_reference_has_term_of_consideration_comment
                ws9['N' + str(7 + i)] = object.archive_reference_has_notif_consider_result_comment
                ws9['O' + str(7 + i)] = object.archive_reference_has_causes_of_failure_comment
                ws9['P' + str(7 + i)] = object.archive_reference_has_sample_document_comment
                ws9['Q' + str(7 + i)] = object.archive_reference_has_document_template_comment
                ws10['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(month, year)
                ws10['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Утверждение схемы расположения земельного участка или земельных участков на кадастровом плане территории" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(month, year)
                ws10['C' + str(7 + i)] = object.land_schemes_id_RGMU
                ws10['D' + str(7 + i)] = object.land_schemes_statement_amount
                ws10['E' + str(7 + i)] = object.land_schemes_link
                ws10['F' + str(7 + i)] = object.land_schemes_has_advanced_appointment_comment
                ws10['G' + str(7 + i)] = object.land_schemes_has_btn_get_service_comment
                ws10['H' + str(7 + i)] = object.land_schemes_has_reglament_comment
                ws10['I' + str(7 + i)] = object.land_schemes_has_estimation_quality_comment
                ws10['J' + str(7 + i)] = object.land_schemes_connected_to_FGIS_DO_comment
                ws10['K' + str(7 + i)] = object.land_schemes_has_electronic_form_printing_comment
                ws10['L' + str(7 + i)] = object.land_schemes_has_edition_draft_comment
                ws10['M' + str(7 + i)] = object.land_schemes_has_term_of_consideration_comment
                ws10['N' + str(7 + i)] = object.land_schemes_has_notif_consider_result_comment
                ws10['O' + str(7 + i)] = object.land_schemes_has_causes_of_failure_comment
                ws10['P' + str(7 + i)] = object.land_schemes_has_sample_document_comment
                ws10['Q' + str(7 + i)] = object.land_schemes_has_document_template_comment

        response = HttpResponse(save_virtual_workbook(wb),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}-{}_Examination_Form.xlsx"'.format(year,month)
        return response

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def export_with_troubles(request,year,month):
    if request.method == 'GET':
        wb = load_workbook('./app/static/app/with_troubles.xlsx')
        thin_border = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))
        objects = get_with_troubles(month, year)
        ws1 = wb["Лист 1"]
        ws2 = wb["Лист 2"]
        ws3 = wb["Лист 3"]
        ws4 = wb["Лист 4"]
        ws5 = wb["Лист 5"]
        ws6 = wb["Лист 6"]
        ws7 = wb["Лист 7"]
        ws8 = wb["Лист 8"]
        ws9 = wb["Лист 9"]
        ws10 = wb["Лист 10"]
        for i in range(len(objects['residential_premises'])):
            object = objects['residential_premises'][i]
            ws1['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Прием заявлений и выдача документов о согласовании проведения переустройства и (или) перепланировки жилого помещения " в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws1['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws1['A' + str(7 + i)] = i + 1
            ws1['B' + str(7 + i)] = object.region_name
            ws1['C' + str(7 + i)] = object.residential_premises_id_RGMU
            ws1['D' + str(7 + i)] = object.residential_premises_statement_amount
            ws1['E' + str(7 + i)] = object.residential_premises_link
            ws1['F' + str(7 + i)] = object.residential_premises_has_advanced_appointment_comment
            ws1['G' + str(7 + i)] = object.residential_premises_has_btn_get_service_comment
            ws1['H' + str(7 + i)] = object.residential_premises_has_reglament_comment
            ws1['I' + str(7 + i)] = object.residential_premises_has_estimation_quality_comment
            ws1['J' + str(7 + i)] = object.residential_premises_connected_to_FGIS_DO_comment
            ws1['K' + str(7 + i)] = object.residential_premises_has_electronic_form_printing_comment
            ws1['L' + str(7 + i)] = object.residential_premises_has_edition_draft_comment
            ws1['M' + str(7 + i)] = object.residential_premises_has_term_of_consideration_comment
            ws1['N' + str(7 + i)] = object.residential_premises_has_notif_consider_result_comment
            ws1['O' + str(7 + i)] = object.residential_premises_has_causes_of_failure_comment
            ws1['P' + str(7 + i)] = object.residential_premises_has_sample_document_comment
            ws1['Q' + str(7 + i)] = object.residential_premises_has_document_template_comment
        for i in range(len(objects['housing_transfer'])):
            object = objects['housing_transfer'][i]
            ws2['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Принятие решений о переводе жилых помещений в нежилые помещения и нежилых помещений в жилые помещения" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month,year)
            ws2['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws2['A' + str(7+i)] = i+1
            ws2['B' + str(7+i)] = object.region_name
            ws2['C' + str(7 + i)] = object.housing_transfer_id_RGMU
            ws2['D' + str(7 + i)] = object.housing_transfer_statement_amount
            ws2['E' + str(7 + i)] = object.housing_transfer_link
            ws2['F' + str(7 + i)] = object.housing_transfer_has_advanced_appointment_comment
            ws2['G' + str(7 + i)] = object.housing_transfer_has_btn_get_service_comment
            ws2['H' + str(7 + i)] = object.housing_transfer_has_reglament_comment
            ws2['I' + str(7 + i)] = object.housing_transfer_has_estimation_quality_comment
            ws2['J' + str(7 + i)] = object.housing_transfer_connected_to_FGIS_DO_comment
            ws2['K' + str(7 + i)] = object.housing_transfer_has_electronic_form_printing_comment
            ws2['L' + str(7 + i)] = object.housing_transfer_has_edition_draft_comment
            ws2['M' + str(7 + i)] = object.housing_transfer_has_term_of_consideration_comment
            ws2['N' + str(7 + i)] = object.housing_transfer_has_notif_consider_result_comment
            ws2['O' + str(7 + i)] = object.housing_transfer_has_causes_of_failure_comment
            ws2['P' + str(7 + i)] = object.housing_transfer_has_sample_document_comment
            ws2['Q' + str(7 + i)] = object.housing_transfer_has_document_template_comment
        for i in range(len(objects['advertising_structures'])):
            object = objects['advertising_structures'][i]
            ws3['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Выдача разрешения на установку и эксплуатацию рекламной конструкции " в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format( month, year)
            ws3['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws3['A' + str(7 + i)] = i + 1
            ws3['B' + str(7 + i)] = object.region_name
            ws3['C' + str(7 + i)] = object.advertising_structures_id_RGMU
            ws3['D' + str(7 + i)] = object.advertising_structures_statement_amount
            ws3['E' + str(7 + i)] = object.advertising_structures_link
            ws3['F' + str(7 + i)] = object.advertising_structures_has_advanced_appointment_comment
            ws3['G' + str(7 + i)] = object.advertising_structures_has_btn_get_service_comment
            ws3['H' + str(7 + i)] = object.advertising_structures_has_reglament_comment
            ws3['I' + str(7 + i)] = object.advertising_structures_has_estimation_quality_comment
            ws3['J' + str(7 + i)] = object.advertising_structures_connected_to_FGIS_DO_comment
            ws3['K' + str(7 + i)] = object.advertising_structures_has_electronic_form_printing_comment
            ws3['L' + str(7 + i)] = object.advertising_structures_has_edition_draft_comment
            ws3['M' + str(7 + i)] = object.advertising_structures_has_term_of_consideration_comment
            ws3['N' + str(7 + i)] = object.advertising_structures_has_notif_consider_result_comment
            ws3['O' + str(7 + i)] = object.advertising_structures_has_causes_of_failure_comment
            ws3['P' + str(7 + i)] = object.advertising_structures_has_sample_document_comment
            ws3['Q' + str(7 + i)] = object.advertising_structures_has_document_template_comment
        for i in range(len(objects['capital_construction'])):
            object = objects['capital_construction'][i]
            ws4['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Выдача разрешения на строительство, реконструкцию объектов капитального строительства" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws4['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws4['A' + str(7 + i)] = i + 1
            ws4['B' + str(7 + i)] = object.region_name
            ws4['C' + str(7 + i)] = object.capital_construction_id_RGMU
            ws4['D' + str(7 + i)] = object.capital_construction_statement_amount
            ws4['E' + str(7 + i)] = object.capital_construction_link
            ws4['F' + str(7 + i)] = object.capital_construction_has_advanced_appointment_comment
            ws4['G' + str(7 + i)] = object.capital_construction_has_btn_get_service_comment
            ws4['H' + str(7 + i)] = object.capital_construction_has_reglament_comment
            ws4['I' + str(7 + i)] = object.capital_construction_has_estimation_quality_comment
            ws4['J' + str(7 + i)] = object.capital_construction_connected_to_FGIS_DO_comment
            ws4['K' + str(7 + i)] = object.capital_construction_has_electronic_form_printing_comment
            ws4['L' + str(7 + i)] = object.capital_construction_has_edition_draft_comment
            ws4['M' + str(7 + i)] = object.capital_construction_has_term_of_consideration_comment
            ws4['N' + str(7 + i)] = object.capital_construction_has_notif_consider_result_comment
            ws4['O' + str(7 + i)] = object.capital_construction_has_causes_of_failure_comment
            ws4['P' + str(7 + i)] = object.capital_construction_has_sample_document_comment
            ws4['Q' + str(7 + i)] = object.capital_construction_has_document_template_comment
        for i in range(len(objects['preschool_education'])):
            object = objects['preschool_education'][i]
            ws5['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Прием заявлений, постановка на учет и зачисление детей в образовательные учреждения, реализующие основную образовательную программу дошкольного образования (детские сады)" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws5['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws5['A' + str(7 + i)] = i + 1
            ws5['B' + str(7 + i)] = object.region_name
            ws5['C' + str(7 + i)] = object.preschool_education_id_RGMU
            ws5['D' + str(7 + i)] = object.preschool_education_statement_amount
            ws5['E' + str(7 + i)] = object.preschool_education_link
            ws5['F' + str(7 + i)] = object.preschool_education_has_advanced_appointment_comment
            ws5['G' + str(7 + i)] = object.preschool_education_has_btn_get_service_comment
            ws5['H' + str(7 + i)] = object.preschool_education_has_reglament_comment
            ws5['I' + str(7 + i)] = object.preschool_education_has_estimation_quality_comment
            ws5['J' + str(7 + i)] = object.preschool_education_connected_to_FGIS_DO_comment
            ws5['K' + str(7 + i)] = object.preschool_education_has_electronic_form_printing_comment
            ws5['L' + str(7 + i)] = object.preschool_education_has_edition_draft_comment
            ws5['M' + str(7 + i)] = object.preschool_education_has_term_of_consideration_comment
            ws5['N' + str(7 + i)] = object.preschool_education_has_notif_consider_result_comment
            ws5['O' + str(7 + i)] = object.preschool_education_has_causes_of_failure_comment
            ws5['P' + str(7 + i)] = object.preschool_education_has_sample_document_comment
            ws5['Q' + str(7 + i)] = object.preschool_education_has_document_template_comment
        for i in range(len(objects['school_education'])):
            object = objects['school_education'][i]
            ws6['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Зачисление детей в общеобразовательные учреждения субъектов Российской Федерации или муниципальные общеобразовательные учреждения" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws6['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws6['A' + str(7 + i)] = i + 1
            ws6['B' + str(7 + i)] = object.region_name
            ws6['C' + str(7 + i)] = object.school_education_id_RGMU
            ws6['D' + str(7 + i)] = object.school_education_statement_amount
            ws6['E' + str(7 + i)] = object.school_education_link
            ws6['F' + str(7 + i)] = object.school_education_has_advanced_appointment_comment
            ws6['G' + str(7 + i)] = object.school_education_has_btn_get_service_comment
            ws6['H' + str(7 + i)] = object.school_education_has_reglament_comment
            ws6['I' + str(7 + i)] = object.school_education_has_estimation_quality_comment
            ws6['J' + str(7 + i)] = object.school_education_connected_to_FGIS_DO_comment
            ws6['K' + str(7 + i)] = object.school_education_has_electronic_form_printing_comment
            ws6['L' + str(7 + i)] = object.school_education_has_edition_draft_comment
            ws6['M' + str(7 + i)] = object.school_education_has_term_of_consideration_comment
            ws6['N' + str(7 + i)] = object.school_education_has_notif_consider_result_comment
            ws6['O' + str(7 + i)] = object.school_education_has_causes_of_failure_comment
            ws6['P' + str(7 + i)] = object.school_education_has_sample_document_comment
            ws6['Q' + str(7 + i)] = object.school_education_has_document_template_comment
        for i in range(len(objects['needing_premises'])):
            object = objects['needing_premises'][i]
            ws7['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Принятие на учет граждан в качестве нуждающихся в жилых помещениях" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws7['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws7['A' + str(7 + i)] = i + 1
            ws7['B' + str(7 + i)] = object.region_name
            ws7['C' + str(7 + i)] = object.needing_premises_id_RGMU
            ws7['D' + str(7 + i)] = object.needing_premises_statement_amount
            ws7['E' + str(7 + i)] = object.needing_premises_link
            ws7['F' + str(7 + i)] = object.needing_premises_has_advanced_appointment_comment
            ws7['G' + str(7 + i)] = object.needing_premises_has_btn_get_service_comment
            ws7['H' + str(7 + i)] = object.needing_premises_has_reglament_comment
            ws7['I' + str(7 + i)] = object.needing_premises_has_estimation_quality_comment
            ws7['J' + str(7 + i)] = object.needing_premises_connected_to_FGIS_DO_comment
            ws7['K' + str(7 + i)] = object.needing_premises_has_electronic_form_printing_comment
            ws7['L' + str(7 + i)] = object.needing_premises_has_edition_draft_comment
            ws7['M' + str(7 + i)] = object.needing_premises_has_term_of_consideration_comment
            ws7['N' + str(7 + i)] = object.needing_premises_has_notif_consider_result_comment
            ws7['O' + str(7 + i)] = object.needing_premises_has_causes_of_failure_comment
            ws7['P' + str(7 + i)] = object.needing_premises_has_sample_document_comment
            ws7['Q' + str(7 + i)] = object.needing_premises_has_document_template_comment
        for i in range(len(objects['town_planning'])):
            object = objects['town_planning'][i]
            ws8['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Выдача градостроительных планов земельных участков" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws8['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws8['A' + str(7 + i)] = i + 1
            ws8['B' + str(7 + i)] = object.region_name
            ws8['C' + str(7 + i)] = object.town_planning_id_RGMU
            ws8['D' + str(7 + i)] = object.town_planning_statement_amount
            ws8['E' + str(7 + i)] = object.town_planning_link
            ws8['F' + str(7 + i)] = object.town_planning_has_advanced_appointment_comment
            ws8['G' + str(7 + i)] = object.town_planning_has_btn_get_service_comment
            ws8['H' + str(7 + i)] = object.town_planning_has_reglament_comment
            ws8['I' + str(7 + i)] = object.town_planning_has_estimation_quality_comment
            ws8['J' + str(7 + i)] = object.town_planning_connected_to_FGIS_DO_comment
            ws8['K' + str(7 + i)] = object.town_planning_has_electronic_form_printing_comment
            ws8['L' + str(7 + i)] = object.town_planning_has_edition_draft_comment
            ws8['M' + str(7 + i)] = object.town_planning_has_term_of_consideration_comment
            ws8['N' + str(7 + i)] = object.town_planning_has_notif_consider_result_comment
            ws8['O' + str(7 + i)] = object.town_planning_has_causes_of_failure_comment
            ws8['P' + str(7 + i)] = object.town_planning_has_sample_document_comment
            ws8['Q' + str(7 + i)] = object.town_planning_has_document_template_comment
        for i in range(len(objects['archive_reference'])):
            object = objects['archive_reference'][i]
            ws9['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Предоставление архивных справок, архивных копий, архивных выписок, информационных писем, связанных с реализацией законных прав и свобод граждан и исполнением государственными органами и органами местного самоуправления своих полномочий" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws9['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws9['A' + str(7 + i)] = i + 1
            ws9['B' + str(7 + i)] = object.region_name
            ws9['C' + str(7 + i)] = object.archive_reference_id_RGMU
            ws9['D' + str(7 + i)] = object.archive_reference_statement_amount
            ws9['E' + str(7 + i)] = object.archive_reference_link
            ws9['F' + str(7 + i)] = object.archive_reference_has_advanced_appointment_comment
            ws9['G' + str(7 + i)] = object.archive_reference_has_btn_get_service_comment
            ws9['H' + str(7 + i)] = object.archive_reference_has_reglament_comment
            ws9['I' + str(7 + i)] = object.archive_reference_has_estimation_quality_comment
            ws9['J' + str(7 + i)] = object.archive_reference_connected_to_FGIS_DO_comment
            ws9['K' + str(7 + i)] = object.archive_reference_has_electronic_form_printing_comment
            ws9['L' + str(7 + i)] = object.archive_reference_has_edition_draft_comment
            ws9['M' + str(7 + i)] = object.archive_reference_has_term_of_consideration_comment
            ws9['N' + str(7 + i)] = object.archive_reference_has_notif_consider_result_comment
            ws9['O' + str(7 + i)] = object.archive_reference_has_causes_of_failure_comment
            ws9['P' + str(7 + i)] = object.archive_reference_has_sample_document_comment
            ws9['Q' + str(7 + i)] = object.archive_reference_has_document_template_comment
        for i in range(len(objects['land_schemes'])):
            object = objects['land_schemes'][i]
            ws10['A1'] = 'Форма самообследования предоставления государственных и муниципальных услуги "Утверждение схемы расположения земельного участка или земельных участков на кадастровом плане территории" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws10['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws10['A' + str(7 + i)] = i + 1
            ws10['B' + str(7 + i)] = object.region_name
            ws10['C' + str(7 + i)] = object.land_schemes_id_RGMU
            ws10['D' + str(7 + i)] = object.land_schemes_statement_amount
            ws10['E' + str(7 + i)] = object.land_schemes_link
            ws10['F' + str(7 + i)] = object.land_schemes_has_advanced_appointment_comment
            ws10['G' + str(7 + i)] = object.land_schemes_has_btn_get_service_comment
            ws10['H' + str(7 + i)] = object.land_schemes_has_reglament_comment
            ws10['I' + str(7 + i)] = object.land_schemes_has_estimation_quality_comment
            ws10['J' + str(7 + i)] = object.land_schemes_connected_to_FGIS_DO_comment
            ws10['K' + str(7 + i)] = object.land_schemes_has_electronic_form_printing_comment
            ws10['L' + str(7 + i)] = object.land_schemes_has_edition_draft_comment
            ws10['M' + str(7 + i)] = object.land_schemes_has_term_of_consideration_comment
            ws10['N' + str(7 + i)] = object.land_schemes_has_notif_consider_result_comment
            ws10['O' + str(7 + i)] = object.land_schemes_has_causes_of_failure_comment
            ws10['P' + str(7 + i)] = object.land_schemes_has_sample_document_comment
            ws10['Q' + str(7 + i)] = object.land_schemes_has_document_template_comment
        for _row in ws1.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws1.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws2.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws2.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws3.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws3.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws4.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws4.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws5.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws5.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws6.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws6.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws7.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws7.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws8.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws8.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws9.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws9.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws10.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws10.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}-{}_With_Troubles_Form.xlsx"'.format(year, month)
        return response

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def export_not_sent(request,year,month):
    if request.method == 'GET':
        wb = load_workbook('./app/static/app/not_sent.xlsx')
        objects = get_not_sent(month, year)
        ws = wb["Лист1"]
        ws['A1'] = 'Список районов , не заполнивших форму за период {}.{} г.'.format(month,year)
        for i in range(len(objects)):
            ws['A'+str(3+i)] = objects[i]
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}-{}_Not_Sent_Form.xlsx"'.format(year, month)
        return response
